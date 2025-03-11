import pandas as pd
import os
from src.configurar_logs import user_logger, dev_logger
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from config import *

def criar_planilha_saida():
    """Cria a planilha de saída se ela não existir."""
    try:
        if not os.path.exists(ARQUIVO_SAIDA):
            df_saida = pd.DataFrame(columns=COLUNAS_SAIDA)
            df_saida.to_excel(ARQUIVO_SAIDA, index=False, engine="openpyxl")
            print(f"Planilha de saída criada em {ARQUIVO_SAIDA}")
            user_logger.info(f"Planilha de saída criada em {ARQUIVO_SAIDA}")
    except Exception as e:
        user_logger.error(f"Erro ao criar planilha de saída: {e}")
        
def preencher_com_dados_existentes():
    """Preenche a tabela de saída com os dados da planilha de entrada."""
    try:
        # Lendo os dados de entrada
        user_logger.info(f"Lendo os dados de entrada em {ARQUIVO_ENTRADA}")
        df_entrada = pd.read_excel(ARQUIVO_ENTRADA, dtype=str, engine="openpyxl")
        df_saida = df_entrada.copy()

        # Criando colunas ausentes
        user_logger.info("Criando e organizando as colunas")
        for coluna in COLUNAS_SAIDA:
            if coluna not in df_saida.columns:
                df_saida[coluna] = ""

        
        # Copiar os dados da coluna 'DIMENSÕES CAIXA (altura x largura x comprimento cm)'
        if "DIMENSÕES CAIXA (altura x largura x comprimento cm)" in df_entrada.columns:
            df_saida["DIMENSÕES CAIXA"] = df_entrada["DIMENSÕES CAIXA (altura x largura x comprimento cm)"]
            # Remover a coluna original para evitar duplicação
            df_saida.drop(columns=["DIMENSÕES CAIXA (altura x largura x comprimento cm)"], inplace=True)

        # Reorganizar as colunas para que 'DIMENSÕES CAIXA' fique após 'E-MAIL'
        colunas_ordenadas = [
            "CNPJ", "RAZÃO SOCIAL", "NOME FANTASIA", "ENDEREÇO", "CEP",
            "DESCRIÇÃO MATRIZ FILIAL", "TELEFONE + DDD", "E-MAIL",
            "VALOR DO PEDIDO", "DIMENSÕES CAIXA", "PESO DO PRODUTO",
            "TIPO DE SERVIÇO JADLOG", "TIPO DE SERVIÇO CORREIOS",
            "VALOR COTAÇÃO JADLOG", "VALOR COTAÇÃO CORREIOS",
            "PRAZO DE ENTREGA CORREIOS", "Status"
        ]
        df_saida = df_saida[colunas_ordenadas]

        return df_saida
    except Exception as e:
        print(f"Erro ao preencher com dados existentes: {e}")

def formatar_colunas_excel():
    
    # Abrir o arquivo Excel com o openpyxl para formatação
    wb = load_workbook(ARQUIVO_SAIDA)
    ws = wb.active

    # Ajustar a largura das colunas com base no tamanho do conteúdo
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter 

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2  # Ajusta um pouco para não ficar muito justo
        ws.column_dimensions[col_letter].width = adjusted_width

    # Salvar o arquivo formatado
    wb.save(ARQUIVO_SAIDA)
    print(f"Colunas formatadas no arquivo {ARQUIVO_SAIDA}")

def finalizar_planilha():
    """ Atualiza o status das cotações e pinta a menor cotação de verde no arquivo de saída."""

    #  Etapa 1: Carregar os dados com pandas
    df = pd.read_excel(ARQUIVO_SAIDA)
    user_logger.info("Carregando os dados com pandas")

    # Atualizar a coluna "Status" com OK
    user_logger.info("Atualizando a coluna 'Status' com OK")
    df["Status"] = df.apply(
        lambda row: "OK" if (
            (pd.notna(row["VALOR COTAÇÃO JADLOG"]) and row["VALOR COTAÇÃO JADLOG"] != "-") and
            (pd.notna(row["VALOR COTAÇÃO CORREIOS"]) and row["VALOR COTAÇÃO CORREIOS"] != "-")
        ) else row["Status"], 
        axis=1
    )

    # Etapa 2: Salvar primeiro o DataFrame atualizado
    user_logger.info("Salvando primeiro o DataFrame atualizado")
    df.to_excel(ARQUIVO_SAIDA, index=False, engine="openpyxl")

    # Etapa 3: Reabrir o arquivo para aplicar a formatação de cores
    user_logger.info("Reabrindo o arquivo para aplicar a formatação de cores")
    wb = load_workbook(ARQUIVO_SAIDA)
    ws = wb.active

    # Definir a cor de preenchimento verde
    user_logger.info("Definindo a cor de preenchimento verde")
    fill_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    # Encontrar as colunas pelo nome
    coluna_jadlog = df.columns.get_loc('VALOR COTAÇÃO JADLOG') + 1
    coluna_correios = df.columns.get_loc('VALOR COTAÇÃO CORREIOS') + 1

    # Aplicar a formatação de cor baseada nos valores
    for row in range(2, len(df) + 2):  
        cotacao_jadlog = ws.cell(row=row, column=coluna_jadlog).value
        cotacao_correios = ws.cell(row=row, column=coluna_correios).value

        # Tentar converter para float (caso os valores sejam strings)
        try:
            cotacao_jadlog = float(cotacao_jadlog) if cotacao_jadlog not in [None, "-"] else None
        except ValueError:
            cotacao_jadlog = None

        try:
            cotacao_correios = float(cotacao_correios) if cotacao_correios not in [None, "-"] else None
        except ValueError:
            cotacao_correios = None
            

        # Aplicar a cor verde ao valor mais barato ou ao valor válido
        if cotacao_jadlog is not None and cotacao_correios is not None:
            if cotacao_jadlog < cotacao_correios:
                ws.cell(row=row, column=coluna_jadlog).fill = fill_green
            else:
                ws.cell(row=row, column=coluna_correios).fill = fill_green
        elif cotacao_jadlog is None and cotacao_correios is not None:
            ws.cell(row=row, column=coluna_correios).fill = fill_green
        elif cotacao_correios is None and cotacao_jadlog is not None:
            ws.cell(row=row, column=coluna_jadlog).fill = fill_green
        user_logger.info(f"Aplicando a cor verde ao valor mais barato na linha {row}")
        
    # Etapa 4: Salvar novamente com a formatação aplicada
    user_logger.info("Salvando novamente com a formatação aplicada")
    wb.save(ARQUIVO_SAIDA)
    
    # Formatar as colunas do arquivo de saída
    formatar_colunas_excel()